import argparse
import os.path
from enum import Enum

from openpyxl.styles import Alignment
from openpyxl import load_workbook
import pandas as pd


class Divisions(Enum):
    """
    Enum with helpers for divisions.
    """
    OPEN = 1,
    MASTERS = 2,
    SENIORS = 3,
    SUPER_SENIORS = 4,
    VETERANS = 5,
    SUPER_VETERANS = 6

    @staticmethod
    def from_string(val):
        lower = val.lower()
        match lower:
            case 'open':
                return Divisions.OPEN
            case 'masters':
                return Divisions.MASTERS
            case 'seniors':
                return Divisions.SENIORS
            case 'super seniors' | "super_seniors":
                return Divisions.SUPER_SENIORS
            case 'veterans':
                return Divisions.VETERANS
            case 'super veterans' | "super_veterans":
                return Divisions.SUPER_VETERANS
            case _:
                raise f"Unexpected value for division: {val}"

    def __str__(self):
        match self:
            case Divisions.OPEN:
                return 'open'
            case Divisions.MASTERS:
                return 'masters'
            case Divisions.SENIORS:
                return 'seniors'
            case Divisions.SUPER_SENIORS:
                return 'super seniors'
            case Divisions.VETERANS:
                return 'veterans'
            case Divisions.SUPER_VETERANS:
                return 'super veterans'
            case _:
                raise f"Unexpected value for division: {self}"


class Sexes(Enum):
    """
    Enum with helpers for sex
    """
    FEMALE = 1,
    MALE = 2

    @staticmethod
    def from_string(val):
        lower = val.lower()
        match lower:
            case 'm' | 'male':
                return Sexes.MALE
            case 'f' | 'female':
                return Sexes.FEMALE
            case _:
                raise f"Unexpected value for Sex: {val}"

    def __str__(self):
        match self:
            case Sexes.MALE:
                return "male"
            case Sexes.FEMALE:
                return "female"
            case _:
                raise f"Unexpected value for Sex: {self}"


def read_file(filename, options):
    """
    Reads an xlsx input file.  Return a list of rows of data.
    # Expected input format is:
    # (first sheet)
    # | Name | Bib | Age | Team | Time |
    :param filename:
    :param options:
    :return:
    """
    source_workbook = load_workbook(filename)
    first_worksheet = source_workbook.worksheets[0]
    start_row = options.data_start_row
    curr_row = start_row
    rows = []
    while True:
        row = first_worksheet[curr_row]
        if row[0].value is None:
            break
        rows.append(row)
        curr_row += 1
    if len(rows) == 0:
        print(f"Warning.  {filename} appears to have no data in the first sheet at row {start_row}")
    source_workbook.close()
    return rows


def process_options():
    """
    Options are: (parens for default values)
    filename: by position (first), defaults to input.xlsx
    data_start_row (0): data row start, default is 2
    :return: options dictionary
    """
    parser = argparse.ArgumentParser(
        prog='scoreit',
        description='Scores XC results')
    parser.add_argument("filename", default="input.xlsx")
    parser.add_argument("--data_start_row", default=2)
    parser.add_argument("--min_team_score_size", default=5)
    return parser.parse_args()


def process_rows(rows, options):
    """
    Reads result data and keeps track of places.  Expected input format:
        headers = ["Name", "Sex", "Bib", "Age", "Team", "Time", "Division"]
    :param rows: the rows of results in spreadsheet format
    :param options: (might eventually include the minimum team count)
    :return: dict with: {
                        "individual": parsed individual results
                        "team_results": team-only data
                        }
    """
    team_results = {}
    individual_results = []
    overall_place = 1
    for row in rows:
        sex = str(Sexes.from_string(row[1].value))  # Normalize
        team = row[4].value
        division = str(Divisions.from_string(row[6].value))  # Normalize
        result = {
                  "Place": overall_place,
                  "Name": row[0].value,
                  "Sex": sex,
                  "Bib": row[2].value,
                  "Age": row[3].value,
                  "Team": team,
                  "Time": row[5].value,
                  "Division": division,
                  }
        overall_place += 1
        individual_results.append(result)
        # Might need to adjust this based on criteria for team scoring.
        # For now, score anyone with a team name,
        # TBD, only score current team depth, e.g. top 5, top 4, etc.
        if len(team):
            if sex not in team_results:
                team_results[sex] = {}
            if division not in team_results[sex]:
                team_results[sex][division] = {}
            if team not in team_results[sex][division]:
                team_results[sex][division][team] = {}
            team_results[sex][division][team][result['Name']] = result

    return {
        "individual": individual_results,
        "team_results": team_results
    }


def team_score_from_team_place(team_place, sex, division):
    """
    Divisions and points from: https://www.pausatf.org/wp-content/uploads/2020/06/PA-Cross-Country-Rules.pdf
    :param team_place: The overall place for team scoring (non-team scorers in overall results don't count)
    :param sex:
    :param division:
    :return: The points scored based on team place, sex and division
    """
    place_table = {
       # Place
       #      "Open, 40+ and Men 50+"
       #    +-------------|           "50+	Women, 60+	Men"
       #    |         +-------------------------|               "60+ Women, and 70+ and 80+"
       #    |         |       +---------------------------------------------|
       #    |         |       |
        1:  [150,     75,     18],
        2:  [135,     63,     13.5],
        3:  [120.5,   52.5,   12],
        4:  [109.5,   48,     10.5],
        5:  [103.5,   43.5,   9],
        6:  [99,      39,     7.5],
        7:  [94.5,    34.5,   6],
        8:  [90,      31.5,   4.5],
        9:  [85.5,    28.5,   3],
        10: [81,      25.5,   1.5],
        11: [76.5,    22.5],
        12: [72,      21],
        13: [67.5,    19.5],
        14: [64.5,    18],
        15: [61.5,    16.5],
        16: [58.5,    15],
        17: [55.5,    13.5],
        18: [52.5,    12],
        19: [49.5,    10.5],
        20: [46.5,    9],
        21: [45,      7.5],
        22: [43.5,    6],
        23: [42,      4.5],
        24: [40.5,    3],
        25: [39,      1.5],
        26: [37.5],
        27: [36],
        28: [34.5],
        29: [33],
        30: [31.5],
        31: [30],
        32: [28.5],
        33: [27],
        34: [25.5],
        35: [24],
        36: [22.5],
        37: [21],
        38: [19.5],
        39: [18],
        40: [16.5],
        41: [15],
        42: [13.5],
        43: [12],
        44: [10.5],
        45: [9],
        46: [7.5],
        47: [6],
        48: [4.5],
        49: [3],
        50: [1.5]
    }
    e_sex = Sexes.from_string(sex)
    e_division = Divisions.from_string(division)
    match e_sex, e_division:
        case Sexes.MALE, Divisions.OPEN | Divisions.MASTERS | Divisions.SENIORS:  # All men under 60
            table_index = 0
        case Sexes.MALE, Divisions.SUPER_SENIORS:  # Men 60-69
            table_index = 1
        case Sexes.MALE, Divisions.VETERANS | Divisions.SUPER_VETERANS:  # Men 70+
            table_index = 2
        case Sexes.FEMALE, Divisions.OPEN | Divisions.MASTERS:  # All women under 50
            table_index = 0
        case Sexes.FEMALE, Divisions.SENIORS:  # Women 50-59
            table_index = 1
        case Sexes.FEMALE, Divisions.SUPER_SENIORS | Divisions.VETERANS | Divisions.SUPER_VETERANS:  # Women 60+
            table_index = 2
        case _:
            raise f"Unknown sex or division Detected: {sex},{division}"

    if team_place not in place_table:
        return 0
    return place_table[team_place][table_index]


def get_race_team_competitors(teams_list):
    """
    Goes through the teams organized lists and generates a list ordered by teams-only placing
    :param teams_list:
    :return: the list of all competitors competing in teams (in place order)
    """
    entrants = []
    for team in teams_list:
        for name in teams_list[team]:
            entrants.append(teams_list[team][name])
    ordered_competitors = sorted(entrants, key=lambda entrant: entrant['Place'])
    return ordered_competitors


def score_team_results(team_results, min_scoring_members):
    """
    Go through the teams results and score the ponts for each team.
    Also sollect stats for each team for displaying to make confirmation and tracking easier.
    :param team_results:
    :param min_scoring_members: TBD
    :return: The scored results.
    """
    scored_results = {}
    for sex in team_results:
        for division in team_results[sex]:
            if sex not in scored_results:
                scored_results[sex] = {}
            if division not in scored_results[sex]:
                scored_results[sex][division] = {
                    "by_team": {}
                }
            scorers = get_race_team_competitors(team_results[sex][division])
            scored_results[sex][division]['overall'] = scorers
            for teams_place_index in range(0, len(scorers)):
                place = teams_place_index+1
                runner_team_points = team_score_from_team_place(place, sex, division)
                scorer = scorers[teams_place_index]
                runner_team = scorer['Team']
                scorer_name = scorer['Name']
                team_results[sex][division][runner_team][scorer_name]['race_teams_place'] = place
                team_results[sex][division][runner_team][scorer_name]['race_teams_points'] = runner_team_points
                if runner_team not in scored_results[sex][division]["by_team"]:
                    scored_results[sex][division]["by_team"][runner_team] = {
                        "runners": [],
                        "score": 0
                    }
                scored_results[sex][division]["by_team"][runner_team]["runners"].append(scorer)
                scored_results[sex][division]["by_team"][runner_team]["score"] += runner_team_points
    return scored_results


def create_writer(filename):
    """
    Create an xlsx writer for a given filename. It will overwrite existing file if present.
    :param filename:
    :return: the writer
    """
    new_file = f"{os.path.splitext(filename)[0]}_processed.xlsx"
    if os.path.exists(new_file):
        exception = None
        try:
            os.remove(new_file)
        except PermissionError as e:
            exception = e
        if exception:
            raise Exception(f"\n\nCan't access {new_file}.  Be sure it is not open in another application.")
    return pd.ExcelWriter(new_file, engine='openpyxl')



def score_race(original_data, options):
    """
    Scores a race and returns a dict with:
    {
        "total_results" : list with the following data:
            # | Place | Name | Bib | Age | Team | Time | Team Points |
        "team_results": ordered list of the following
            # | Team Name | Total Team Points | Teams Member Count |
        }
    }

    :param original_data:
    :param options: contains like:
        data start row
    :return: workbook, results dict
    """

    processed_rows = process_rows(original_data, options)
    scored_team_results = score_team_results(processed_rows["team_results"], options.min_team_score_size)

    return {
        "total_results": processed_rows,
        "team_results": scored_team_results
    }


def write_summary(workbook, results_sheet, results):
    """
    :param workbook: the current workbook
    :param results_sheet: the sheeet to write the results to, starting at row 0
    :param results: the results to write
    :return: the last row written to
    """
    # Write header
    headers = ["Team Name", "Total Team Points", "Teams Member Count"]
    col = 0
    row = 0
    bold_format = workbook.add_format({'bold': True})
    for header in headers:
        results_sheet.write(0, col, header, bold_format)
        col += 1

    row += 1
    for team_result in results["team_results"]:
        col = 0
        for result in team_result:
            results_sheet.write(result)
            col += 1
        row += 1
    return row


# def write_calculations(workbook, sheet, results, start_row):
#     row = start_row
#     headers = [" Place", "Name", "Bib", "Age", "Team", "Time", "Team Points"]
#
#     bold_format = workbook.add_format({'bold': True})
#     col = 0
#     for header in headers:
#         sheet.write(0, col, header, bold_format)
#         col += 1
#
#     row += 1
#     for result in results["total_results"]:
#         for col in result:
#             sheet.write(row, col, result[col])
#             col += 1
#         row += 1


def format_team_info(results_by_team):
    """
    Create a dict with results organized by team for displaying team scoring results
    :param results_by_team:
    :return: the restructured based on team place
    """
    reformatted_results_by_team = []
    for team in results_by_team:
        team_info = ""
        for runner in results_by_team[team]['runners']:
            place = runner['Place']
            name = runner['Name']
            teams_place = runner['race_teams_place']
            teams_points = runner['race_teams_points']
            team_info += f"{teams_points:.1f} {place:3}  {teams_place:3} {name:3}\n"
        team_info = team_info.rstrip("\n")
        reformatted_results_by_team.append({
            'score': results_by_team[team]['score'],
            'team': team,
            'runners': team_info
        })
    dict_results = {}
    place = 1
    for i in reformatted_results_by_team:
        dict_results[place] = i
        place += 1
    return dict_results


def get_width_for_col(sheet, col_number):
    """
    Gets the largest required column width for a given column
    :param sheet:
    :param col_number:
    :return:
    """
    col = list(sheet.columns)[col_number]
    max_width = 10
    for cell in col:
        # determine max width
        value = cell.value
        if None is value or not isinstance(value, str):
            continue
        lines = cell.value.split('\n')
        for line in lines:
            width = len(line)
            max_width = max(width, max_width)
    return max_width


def get_height_for_row(sheet, row_number):
    """
    Gets the largest required row height for a given row
    :param sheet:
    :param row_number:
    :return:
    """
    row = list(sheet.rows)[row_number]
    height = 14
    for cell in row:
        value = cell.value
        if None is value or not isinstance(value, str):
            continue
        lines = 1 + cell.value.count("\n")
        height = max(height, lines * height)
    return height


def left_align_sheet_data(sheet):
    """
    Go through an entire sheet and format everyting to be left aligned
    :param sheet:
    """
    for row in sheet.rows:
        for cell in list(row):
            cell.alignment = Alignment(horizontal='left')


def write_results_sheets(results, options):
    """
    Write all the results.
    The first sheet contains the individual results.
    Subsequent sheets include overall team results and then the same formatted by team with totals.
    :param results:
    :param options:
    """
    writer = create_writer(options.filename)

    df_individual = pd.DataFrame.from_dict(results['total_results']['individual'])
    df_individual.to_excel(writer, sheet_name="individual")
    for sex in results['team_results']:
        for division in results['team_results'][sex]:
            race_results = results['team_results'][sex][division]
            df_overall = pd.DataFrame.from_dict(race_results['overall'])
            curr_sheet_name = f"{sex}-{division}-overall"
            df_overall.to_excel(writer, sheet_name=curr_sheet_name)
            curr_sheet = writer.sheets[curr_sheet_name]
            for column in curr_sheet.columns:
                curr_sheet.column_dimensions[column[0].column_letter].auto_size = True
            left_align_sheet_data(curr_sheet)

            minimal_by_team = format_team_info(race_results['by_team'])
            df_team = pd.DataFrame.from_dict(minimal_by_team)
            curr_sheet_name = f"{sex}-{division}-by_team"
            df_team.to_excel(writer, sheet_name=curr_sheet_name)
            curr_sheet = writer.sheets[curr_sheet_name]
            curr_sheet_columns = list(curr_sheet.columns)
            for i in range(0, curr_sheet.max_column):
                curr_col_letter = curr_sheet_columns[i][0].column_letter
                curr_sheet.column_dimensions[curr_col_letter].width = get_width_for_col(curr_sheet, i)
            for i in range(0, curr_sheet.max_row):
                desired_row_height = int(get_height_for_row(curr_sheet, i))
                curr_sheet.row_dimensions[i+1].height = desired_row_height
            left_align_sheet_data(curr_sheet)
    writer.close()


def main():
    """
    Given some cmdline options, reads an xlsx results file, processes the data and then creates a new results xlsx file.
    """
    # Process any passed in options
    options = process_options()
    # Read the source xlsx file
    original_data = read_file(options.filename, options)
    # Score the race
    results = score_race(original_data, options)
    # Write the results out to a new file, replacing any existing file.
    write_results_sheets(results, options)


if __name__ == "__main__":
    main()
